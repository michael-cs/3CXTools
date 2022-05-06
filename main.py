import PySimpleGUI as sg
import csv
import xml.etree.ElementTree as et
import openpyxl
from random_functions import *

def tela_inicial():
    sg.change_look_and_feel('DarkBlue16')
    FILENAME = r'img_1.png'
    DISPLAY_TIME_MILLISECONDS = 4000
    sg.Window('Window Title', [[sg.Image(FILENAME)]], transparent_color=sg.theme_background_color(), no_titlebar=True, keep_on_top=True).read(timeout=DISPLAY_TIME_MILLISECONDS, close=True)
    FILENAME = r'img_2.png'
    tela_inicio_layout = [
        [sg.Image(FILENAME)],
        [sg.Text('Selecione a ferramenta desejada: ')],
        [sg.Text('___________________________________________________________')],
        [sg.Text('          Extrair Contatos Pessoais:___________________'),sg.Button('Ok',key="select_cont")],
        [sg.Text('          Criar Extensões: ___________________________'),sg.Button('Ok',key="select_ext")],
        [sg.Text('          Criar Regras de Entrada:_____________________'),sg.Button('Ok',key="select_reg")],
        [sg.Button('Sair')]
    ] 
    
    tela_inicio=sg.Window("3CX Tools - By Michatec").layout(tela_inicio_layout)
    
    while True:
        event,values = tela_inicio.Read()
        if event in ( 'Sair',sg.WIN_CLOSED):
            print('closing...')
            break
        elif event == 'select_cont':
            tela_contatos() 
        elif event == 'select_ext':
            tela_extensoes() 
        elif event == 'select_reg':
            tela_regras() 

    tela_inicio.close()            


def tela_contatos():
    sg.change_look_and_feel('DarkBlue16')
    FILENAME = r'img_2.png'
   
    tela_cont_layout = [
        [sg.Text("    "),sg.Image(FILENAME)],
        [sg.Text('3CX backup xml archive: '),sg.Input(key='xml')],
        [sg.Text('Extension to extract PB: '), sg.Input(key='extension')],
        [sg.Button('Ok')]
    ] 
    
    tela_cont=sg.Window("Personal Phonebook Extractor - By Michatec").layout(tela_cont_layout)
       
    while True:
        event,values = tela_cont.Read()
        if event in (None,sg.WIN_CLOSED):
            break
        elif event == 'Ok':
            xml3cx=values['xml']
            extension3cx=values['extension']
            if xml3cx == "" or extension3cx == "":
                print("Sem Infos")
            else:
                extrair_contatos(xml3cx,extension3cx)
            break          

    tela_cont.close()

      
def tela_extensoes():

    sg.change_look_and_feel('DarkBlue16')
    FILENAME = r'img_2.png'

    tela_ext_layout = [
        [sg.Image(FILENAME)],
        [sg.Text("          "),sg.Text('Planilha de Pré-Instalação:')],
        [sg.Text("          "),sg.Input(key='planilha')],
        [sg.Text("          "),sg.Checkbox('Enviar credenciais para email específico?', default=False, key='checkbox1')],
        [sg.Text("          "),sg.Input(key='email')],
        [sg.Button('Ok')]
    ]
    tela_ext = sg.Window('3CX Extension Creator - By Michatec').layout(tela_ext_layout)

    while True:
        event,values = tela_ext.read()  
        if event in (None,sg.WIN_CLOSED):
            break
        elif event == 'Ok':           
            planilha_pre=values['planilha']
            email_envio=values['email']
            checkbox_email=values['checkbox1']
            if planilha_pre == "" or email_envio == "":
                print("Sem Infos")
            else:
                cria_extensoes(planilha_pre,checkbox_email,email_envio)           
            break 
    tela_ext.close()


def tela_regras():
    sg.change_look_and_feel('DarkBlue16')
    FILENAME = r'img_2.png'

    tela_reg_layout = [
    [sg.Text("         "),sg.Image(FILENAME)],
    [sg.Text('Indice do tronco(Máscara utilzada pelo sistema 3CX.Ex:10000,10001...):'), sg.Input(key='indicetronco',size=10)],
    [sg.Text('DDR Inicial:'), sg.Input(key='DDR inicial',size=12), sg.Text('DDR Final'), sg.Input(key='DDR final',size=12)],
    [sg.Text('_____________________________Diurno_______________________________________')],
    [sg.Checkbox('Enviar todos os DIDs para o mesmo destino.', default=False, key='checkdestinodiu')],               
    [sg.Text('Primeiro Destino:'), sg.Input(key='primeirodestinodiu',size=10), sg.Text('Tipo de destino:'), sg.Checkbox('Ramal',default=False, key='checktiporamaldiu'), sg.Checkbox('URA',default=False,key='checktipouradiu')],
    [sg.Text('____________________________Noturno_______________________________________')],
    [sg.Checkbox('Enviar todos os DIDs para o mesmo destino.', default=False, key='checkdestinonot')],               
    [sg.Text('Primeiro Destino:'), sg.Input(key='primeirodestinonot',size=10), sg.Text('Tipo de destino:'), sg.Checkbox('Ramal',default=False, key='checktiporamalnot'), sg.Checkbox('URA',default=False,key='checktipouranot')],
    [sg.Button('Ok')]
    ]
    
    tela_reg = sg.Window('3CX Inbound Rules Creator - By Michatec').layout(tela_reg_layout)
    while True:
        event,values = tela_reg.read()  
        if event in (None,sg.WIN_CLOSED):
            break
        elif event == 'Ok':           
            index_trunk=values['indicetronco']
            n_ddr_inicial=values['DDR inicial']
            n_ddr_final=values['DDR final']

            destino_unico_diu=values['checkdestinodiu']
            pri_destino_diu=values['primeirodestinodiu']
            tipo_diu_ramal=values['checktiporamaldiu']
            tipo_diu_ura=values['checktipouradiu']

            destino_unico_not=values['checkdestinonot']
            pri_destino_not=values['primeirodestinonot']
            tipo_not_ramal=values['checktiporamalnot']
            tipo_not_ura=values['checktipouranot']
            if index_trunk == "" or n_ddr_inicial == "" or n_ddr_final == "" or pri_destino_diu == "" or pri_destino_not =="":
                print("Sem Infos") 
            else:
                cria_regras_entrada(index_trunk,n_ddr_inicial,n_ddr_final,destino_unico_diu,destino_unico_not,pri_destino_diu,pri_destino_not,tipo_diu_ramal,tipo_diu_ura,tipo_not_ramal,tipo_not_ura)
            break 
    tela_reg.close()


def cria_extensoes(planilha_pre,checkbox_email,email_envio):
    name=lastname=number=email=data=''
    f = open('extensions.csv','w', encoding='UTF8',newline='')
    cabecalho = ['Number','FirstName','LastName','EmailAddress','MobileNumber','AuthID','AuthPassword','WebMeetingFriendlyName','WebMeetingPrivateRoom','ClickToCall','ClickToCallFriendlyName','WebMeetingAcceptReject','EnableVoicemail','VMNoPin','VMPlayCallerID','PIN','VMPlayMsgDateTime','VMEmailOptions','QueueStatus','OutboundCallerID','SIPID','DeliverAudio','SupportReinvite','SupportReplaces','EnableSRTP','ManagementAccess','ReporterAccess','WallboardAccess','TurnOffMyPhone','HideFWrules','CanSeeRecordings','CanDeleteRecordings','RecordCalls','CallScreening','EmailMissedCalls','Disabled','DisableExternalCalls','AllowLanOnly','BlockRemoteTunnel','PinProtect','MAC_0','InterfaceIP_0','UseTunnel','DND','UseCTI','StartupScreen','HotelModuleAccess','DontShowExtInPHBK','DeskphoneWebPass','SrvcAccessPwd','VoipAdmin','SysAdmin','SecureSIP','PhoneModel14','PhoneTemplate14','CustomTemplate','PhoneSettings','AllowAllRecordings','PushExtension','Integration','AllowOwnRecordings','RecordExternalCallsOnly','DID','SMS','PhoneSysAdmin']
    writer = csv.writer(f)
    writer.writerow(cabecalho)  

    try:
        path = planilha_pre
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active
        m_row=sheet_obj.max_row
        max_col = sheet_obj.max_column

    except TypeError:
        print('Valores inválidos!')

    try:
        for line in range (6, m_row+1):
            for i in range(7, 11):
                cell_obj = sheet_obj.cell(row = line, column = i)         
                if i ==7:
                    number = cell_obj.value          
                elif i==8:
                    name=cell_obj.value
                elif i==9:
                    lastname=cell_obj.value
                elif i==10:            
                    email=cell_obj.value
                    if number!=None:
                        senha = pass_random()
                        meeting_address= meeting_random()
                        pin_number=pin_random()
                        if checkbox_email == True:
                            email = email_envio
                        data=[number,name,lastname,email,'',senha,senha,meeting_address,0,0,'',0,0,1,0,pin_number,0,0,1,'','',0,1,1,0,0,0,'',0,0,0,0,0,0,0,0,0,0,0,0,'','',1,0,0,0,'',0,senha,senha,0,0,0,'','','','',0,1,'',1,0,'','',0]   
                        with open('extensions.csv', 'a', encoding='UTF8',newline='') as f:
                            writer = csv.writer(f)                                                                                          
                            writer.writerow(data)
                        print(data)
    except:
        print('Encerrando...')


def extrair_contatos(xml3cx, extension3cx):
    tree = et.ElementTree(file=xml3cx)
    root = tree.getroot()
    name=lastname=company=phone1=phone2=home1=home2=number=number2=email=other=fax1=fax2=test=''
    f = open('contacts.csv','w')
    header = ['FirstName','LastName','Company','Mobile','Mobile2','Home','Home2','Business','Business2','Email','Other','BusinessFax','HomeFax','Pager']
    writer = csv.writer(f)
    writer.writerow(header)  

    try:
        for globals in root:
            #print(globals.tag)
            for tenants in globals:
                #print(tenants.tag,tenants.attrib)
                for tenant in tenants:
                    #print(tenant.tag,tenant.attrib)
                    for dn in tenant:
                        #print(dn.tag,dn.attrib)
                        for extension in dn:
                            #print(extension.tag,extension.text)
                            if extension.tag == 'Number' and extension.text == extension3cx:
                                #print(extension.tag,extension.text)
                                for extension in dn:
                                    #print(extension.tag,extension.text)
                                    for phonebookentries in extension:
                                        if phonebookentries.tag == 'PhoneBookEntry':
                                            #print(phonebookentries.tag,phonebookentries.text,phonebookentries.attrib)
                                            for phonebookentry in phonebookentries:                                        
                                                contact=[phonebookentry.tag,phonebookentry.text]                                                                  
                                                if contact[0]=='FirstName':
                                                    name=contact[1]                              
                                                elif contact[0]=='PhoneNumber':
                                                    phone1=contact[1]                                                                           
                                                elif contact[0]=='LastName':
                                                    lastname=contact[1]
                                                elif contact[0]=='CompanyName':
                                                    company=contact[1]
                                                elif contact[0]=='AddressNumberOrData0':
                                                    phone2=contact[1]
                                                elif contact[0]=='AddressNumberOrData1':
                                                    home1=contact[1]
                                                elif contact[0]=='AddressNumberOrData2':
                                                    home2=contact[1]
                                                elif contact[0]=='AddressNumberOrData3':
                                                    number=contact[1]
                                                elif contact[0]=='AddressNumberOrData4':
                                                    number2=contact[1]
                                                elif contact[0]=='AddressNumberOrData5':
                                                    email=contact[1]
                                                elif contact[0]=='AddressNumberOrData6':
                                                    other=contact[1]
                                                elif contact[0]=='AddressNumberOrData7':
                                                    fax1=contact[1]
                                                elif contact[0]=='AddressNumberOrData8':
                                                    fax2=contact[1]                                                             
                                            data = [name,lastname,company,phone1,phone2,home1,home2,number,number2,email,other,fax1,fax2,'']
                                            name=lastname=company=phone1=phone2=home1=home2=number=number2=email=other=fax1=fax2=test=''
                                            print(data)
                                            with open('contacts.csv', 'a', encoding='UTF8', newline='') as f:
                                                writer = csv.writer(f)                                                                                          
                                                writer.writerow(data)
    except:
        print("error")


def cria_regras_entrada(index_trunk,n_ddr_inicial,n_ddr_final,destino_unico_diu,destino_unico_not,pri_destino_diu,pri_destino_not,tipo_diu_ramal,tipo_diu_ura,tipo_not_ramal,tipo_not_ura):
    dest_diu=(int(pri_destino_diu)-1)
    ddr_inicial=int(n_ddr_inicial)
    ddr_final=int(n_ddr_final)
    dest_not=(int(pri_destino_not)-1)
    trunk_id=[int(index_trunk)]
    diu_type=''
    notu_type=''

    f = open('exportDID.csv','w', encoding='UTF8',newline='')
    cabecalho = ['PRIORITY','NAME','TYPE','MASK','PORTS','INOFFICE_DEST_TYPE','INOFFICE_DEST_NUMBER','SAME_DEST_AS_INOFFICE','SPECIFIC_HOURS','SPECIFIC_HOURS_TIME','INCLUDE_HOLIDAYS','OUTOFOFFICE_DEST_TYPE','OUTOFOFFICE_DEST_NUMBER','PLAY_HOLIDAY_PROMPT']
    writer = csv.writer(f)
    writer.writerow(cabecalho)

    for ddr in range(ddr_inicial,ddr_final+1):
        ddr+1

        #VALIDAÇÃO DE CHEKBOXES DE DESTINO UNICO
        #DIURNO
        if destino_unico_diu == True:
            dest_diu=pri_destino_diu
        else:
            dest_diu+=1
        #NOTURNO
        if destino_unico_not == True:
            dest_not=pri_destino_not
        else:
            dest_not+=1

        #VALIDAÇÃO DE CHECKBOXES DE TIPOS DE DESTINO
        #DIURNO
        if tipo_diu_ramal == True:
            diu_type=2
        elif tipo_diu_ura == True:
            diu_type=5
        elif tipo_diu_ramal == tipo_diu_ura:
            diu_type=2
        #NOTURNO
        if tipo_not_ramal == True:
            notu_type=2
        elif tipo_not_ura == True:
            notu_type=5
        elif tipo_not_ramal == tipo_not_ura:
            notu_type=2 
                
        data=['','','1',ddr,trunk_id,diu_type,dest_diu,'0','0','','',notu_type,dest_not,'0']   
        with open('exportDID.csv', 'a', encoding='UTF8',newline='') as f:
            writer = csv.writer(f)                                                                                          
            writer.writerow(data)
            print(data)


tela_inicial()